from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
import uuid
import threading
import time
import logging
import re
from datetime import datetime

from .models import Utilisateur, HistoriquePresence, ConfigurationSession, BadgeVisiteur
from .nfc_service import get_nfc_service
from django.contrib.auth.hashers import make_password, check_password
from django.db import transaction, IntegrityError

logger = logging.getLogger(__name__)


# Service NFC global
nfc_service = None

# Note: On ne fait plus de cooldown; on se base uniquement sur la présence continue de la carte

# Mémoire de la carte actuellement posée sur le lecteur (edge detection)
last_seen_uid = None


# --- Rétention des données ---
def _get_cfg(key: str, default: str) -> str:
    """Lire/initialiser une valeur de configuration simple dans ConfigurationSession."""
    try:
        cfg = ConfigurationSession.objects.get(cle=key)
        return cfg.valeur
    except ConfigurationSession.DoesNotExist:
        ConfigurationSession.objects.create(cle=key, valeur=str(default))
        return str(default)


def _set_cfg(key: str, value: str) -> None:
    try:
        cfg = ConfigurationSession.objects.get(cle=key)
        cfg.valeur = str(value)
        cfg.save()
    except ConfigurationSession.DoesNotExist:
        ConfigurationSession.objects.create(cle=key, valeur=str(value))


def _purge_logs_older_than(days: int):
    """Nettoyer logs/app.log en ne conservant que les lignes des N derniers jours.
    Hypothèse: format '%(asctime)s [LEVEL] logger: message'.
    """
    log_file = (settings.BASE_DIR / 'logs' / 'app.log')
    if not log_file.exists():
        return
    cutoff = timezone.now() - timezone.timedelta(days=days)
    # Pattern de début de ligne: 2025-11-05 12:34:56,789
    ts_re = re.compile(r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}) ")
    keep_lines = []
    try:
        with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                m = ts_re.match(line)
                if not m:
                    # Si non parsable, on le conserve pour éviter de perdre des infos utiles
                    keep_lines.append(line)
                    continue
                try:
                    ts = datetime.strptime(m.group('ts'), '%Y-%m-%d %H:%M:%S,%f')
                    ts = timezone.make_aware(ts, timezone.get_current_timezone()) if settings.USE_TZ else ts
                    if ts >= cutoff:
                        keep_lines.append(line)
                except Exception:
                    keep_lines.append(line)
        with open(log_file, 'w', encoding='utf-8') as f:
            f.writelines(keep_lines)
    except Exception as e:
        logger.warning(f"Purge logs échouée: {e}")


def enforce_retention_policies():
    """Appliquer les règles de conservation:
    - Historique (entrées/sorties): 90 jours par défaut (clé retention_history_days)
    - Logs: 90 jours par défaut (clé retention_logs_days)
    - Données utilisateurs: conservées jusqu'à suppression admin (rien à faire ici)
    Cette fonction s'exécute au plus une fois par 24h (clé last_purge_ts).
    """
    try:
        last_purge_str = _get_cfg('last_purge_ts', '')
        should_run = True
        if last_purge_str:
            try:
                last_purge = datetime.fromisoformat(last_purge_str)
                # Comparer en timezone naive, c'est suffisant pour notre but quotidien
                should_run = (datetime.now() - last_purge).total_seconds() > 23*3600
            except Exception:
                should_run = True
        if not should_run:
            return

        # Durées (jours)
        try:
            hist_days = int(_get_cfg('retention_history_days', '90'))
        except ValueError:
            hist_days = 90
        try:
            log_days = int(_get_cfg('retention_logs_days', '90'))
        except ValueError:
            log_days = 90
        # Utilisateurs inactifs (optionnel)
        try:
            users_inactive_enabled = _get_cfg('retention_users_inactive_enabled', '0') in ('1', 'true', 'True')
        except Exception:
            users_inactive_enabled = False
        try:
            users_inactive_days = int(_get_cfg('retention_users_inactive_days', '0'))
        except ValueError:
            users_inactive_days = 0

        # Purge HistoriquePresence
        cutoff_hist = timezone.now() - timezone.timedelta(days=hist_days)
        deleted_count, _ = HistoriquePresence.objects.filter(horodatage__lt=cutoff_hist).delete()
        if deleted_count:
            logger.info(f"Purge historique: {deleted_count} enregistrements < {cutoff_hist}")

        # Purge logs
        _purge_logs_older_than(log_days)

        # Purge utilisateurs inactifs (si activée)
        if users_inactive_enabled and users_inactive_days >= 1:
            cutoff_users = timezone.now() - timezone.timedelta(days=users_inactive_days)
            # Déterminer la dernière activité: dernière entrée/sortie sinon date_creation
            # On évite de supprimer des personnes marquées présentes
            to_delete_ids = []
            from django.db.models import Max
            last_acts = (
                HistoriquePresence.objects
                .values('utilisateur_id')
                .annotate(last=Max('horodatage'))
            )
            last_map = {row['utilisateur_id']: row['last'] for row in last_acts}
            for u in Utilisateur.objects.filter(present=False):
                last_activity = last_map.get(u.id) or u.date_creation
                # last_activity est naïf/aware selon USE_TZ; on compare avec cutoff_users qui est aware
                if settings.USE_TZ and timezone.is_naive(last_activity):
                    last_activity = timezone.make_aware(last_activity, timezone.get_current_timezone())
                if last_activity < cutoff_users:
                    to_delete_ids.append(u.id)
            if to_delete_ids:
                count = Utilisateur.objects.filter(id__in=to_delete_ids).delete()[0]
                logger.info(f"Purge utilisateurs inactifs: {count} supprimés (< {cutoff_users})")

        _set_cfg('last_purge_ts', datetime.now().isoformat())

        # Désaffectation quotidienne des badges visiteurs (au premier appel après 00:00)
        try:
            today_str = datetime.now().date().isoformat()
            last_reset = _get_cfg('last_visitor_reset_date', '')
            if last_reset != today_str:
                # On désaffecte tous les badges visiteurs
                count = BadgeVisiteur.objects.filter(affecte_a__isnull=False).update(affecte_a=None, date_attribution=None)
                if count:
                    logger.info(f"Reset quotidien des badges visiteurs: {count} badge(s) désaffecté(s)")
                _set_cfg('last_visitor_reset_date', today_str)
        except Exception as e:
            logger.warning(f"Reset badges visiteurs échoué: {e}")
    except Exception as e:
        logger.warning(f"Enforce retention échouée: {e}")

def get_admin_status(request):
    """Vérifier si l'utilisateur est en mode admin"""
    return request.session.get('is_admin', False)

def toggle_admin_mode(request):
    """Basculer entre mode utilisateur et admin"""
    if request.method == 'POST':
        if get_admin_status(request):
            # Déconnexion admin
            request.session['is_admin'] = False
            messages.success(request, "Déconnecté du mode administrateur")
        else:
            # Connexion admin via mot de passe stocké en configuration (hashé)
            password = request.POST.get('admin_password', '')
            # Récupérer le mot de passe depuis ConfigurationSession
            try:
                cfg = ConfigurationSession.objects.get(cle='admin_password')
                stored_hash = cfg.valeur
            except ConfigurationSession.DoesNotExist:
                # Initialiser avec admin123 si absent
                stored_hash = make_password('admin123')
                ConfigurationSession.objects.create(cle='admin_password', valeur=stored_hash)
            if check_password(password, stored_hash):
                request.session['is_admin'] = True
                messages.success(request, "Connecté en mode administrateur")
            else:
                messages.error(request, "Mot de passe administrateur incorrect")
    
    return redirect('accueil')

def accueil(request):
    """Page d'accueil - Gestion de présence"""
    global nfc_service
    # Appliquer périodiquement les règles de rétention
    enforce_retention_policies()
    
    # Initialiser le service NFC si ce n'est pas fait
    if nfc_service is None:
        nfc_service = get_nfc_service()
        # Démarrer la surveillance continue automatiquement
        try:
            nfc_service.start_continuous_monitoring()
        except Exception as e:
            logger.warning(f"Impossible de démarrer la surveillance NFC: {e}")
    
    # Compter les personnes présentes
    nombre_presents = Utilisateur.objects.filter(present=True).count()

    # Barre de recherche (dans les présents uniquement)
    search_query = request.GET.get('search', '').strip()
    utilisateurs_presents = Utilisateur.objects.filter(present=True)
    if search_query:
        utilisateurs_presents = utilisateurs_presents.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(societe_raison__icontains=search_query)
        )
    utilisateurs_presents = utilisateurs_presents.order_by('nom', 'prenom')
    
    # Liste des personnes présentes avec leur dernière entrée
    personnes_presentes = []
    
    for utilisateur in utilisateurs_presents:
        # Trouver la dernière entrée
        derniere_entree = HistoriquePresence.objects.filter(
            utilisateur=utilisateur,
            type_action='ENTREE'
        ).order_by('-horodatage').first()
        
        utilisateur.heure_entree = derniere_entree.horodatage if derniere_entree else None
        personnes_presentes.append(utilisateur)
    
    # Vérifier le mode admin
    is_admin = get_admin_status(request)
    
    context = {
        'nombre_presents': nombre_presents,
        'personnes_presentes': personnes_presentes,
        'is_admin': is_admin,
        'search_query': search_query,
    }
    
    return render(request, 'presence/accueil.html', context)

def _resoudre_uid(uid):
    """Résoudre un UID NFC : badge visiteur → utilisateur → inconnu.
    Retourne un dict prêt pour JsonResponse.
    """
    # 1) Badge visiteur
    badge = BadgeVisiteur.objects.select_related('affecte_a').filter(uid_carte__iexact=uid).first()
    if badge is not None:
        if badge.affecte_a is None:
            return {'success': True, 'action': 'visiteur_affectation', 'uid': uid,
                    'reader_connected': True, 'message': 'Badge visiteur non affecté'}
        utilisateur = badge.affecte_a
        if utilisateur.blackliste:
            logger.warning(f"BLACKLIST (visiteur): {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'blacklist', 'reader_connected': True,
                    'message': f"Utilisateur blacklisté: {utilisateur.prenom} {utilisateur.nom}",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}
        if utilisateur.statut_validation == 'EN_ATTENTE':
            logger.warning(f"EN_ATTENTE (visiteur): {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'en_attente', 'reader_connected': True,
                    'message': f"Utilisateur en attente de validation: {utilisateur.prenom} {utilisateur.nom}",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}
        if utilisateur.present:
            utilisateur.present = False
            utilisateur.save()
            HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='SORTIE')
            logger.info(f"SORTIE (visiteur): {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'sortie', 'reader_connected': True,
                    'message': f"{utilisateur.prenom} {utilisateur.nom} a quitté les locaux",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}
        utilisateur.present = True
        utilisateur.save()
        HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='ENTREE')
        logger.info(f"ENTREE (visiteur): {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
        return {'success': True, 'action': 'entree', 'reader_connected': True,
                'message': f"{utilisateur.prenom} {utilisateur.nom} est arrivé",
                'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}

    # 2) Carte personnelle d'un utilisateur
    utilisateur = Utilisateur.objects.filter(uid_carte__iexact=uid).first()
    if utilisateur is not None:
        if utilisateur.blackliste:
            logger.warning(f"BLACKLIST: {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'blacklist', 'reader_connected': True,
                    'message': f"Utilisateur blacklisté: {utilisateur.prenom} {utilisateur.nom}",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'societe_raison': utilisateur.societe_raison}}
        if utilisateur.statut_validation == 'EN_ATTENTE':
            logger.warning(f"EN_ATTENTE: {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'en_attente', 'reader_connected': True,
                    'message': f"Utilisateur en attente de validation: {utilisateur.prenom} {utilisateur.nom}",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'societe_raison': utilisateur.societe_raison}}
        if utilisateur.present:
            utilisateur.present = False
            utilisateur.save()
            HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='SORTIE')
            logger.info(f"SORTIE: {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
            return {'success': True, 'action': 'sortie', 'reader_connected': True,
                    'message': f"{utilisateur.prenom} {utilisateur.nom} a quitté les locaux",
                    'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}
        utilisateur.present = True
        utilisateur.save()
        HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='ENTREE')
        logger.info(f"ENTREE: {utilisateur.prenom} {utilisateur.nom} ({utilisateur.societe_raison}) - UID={uid}")
        return {'success': True, 'action': 'entree', 'reader_connected': True,
                'message': f"{utilisateur.prenom} {utilisateur.nom} est arrivé",
                'utilisateur': {'nom': utilisateur.nom, 'prenom': utilisateur.prenom, 'entreprise': utilisateur.societe_raison}}

    # 3) Carte totalement inconnue
    logger.info(f"CARTE INCONNUE détectée UID={uid}")
    return {'success': True, 'action': 'lier_badge_utilisateur', 'uid': uid,
            'reader_connected': True, 'message': 'Badge inconnu - veuillez le lier à un utilisateur'}


@csrf_exempt
@require_http_methods(["POST"])
def lire_carte_nfc(request):
    """Endpoint pour lire une carte NFC"""
    global nfc_service

    try:
        if nfc_service is None:
            nfc_service = get_nfc_service()
        uid = nfc_service.lire_carte()
        if uid:
            return JsonResponse(_resoudre_uid(uid))
        return JsonResponse({'success': False, 'message': 'Aucune carte détectée'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur de lecture NFC: {str(e)}'})

@csrf_exempt
@require_http_methods(["POST"])
def enregistrer_utilisateur(request):
    """Enregistrer un nouvel utilisateur en attente de validation"""
    try:
        data = json.loads(request.body)
        # Si aucun UID fourni (ex: création depuis un badge visiteur non personnel), on génère un UID fictif unique
        raw_uid = (data.get('uid') or '').strip()
        if not raw_uid:
            # Générer un identifiant synthétique de 32 caractères max (contrainte du modèle)
            # Préfixe 'V' pour marquer un utilisateur créé sans carte personnelle
            raw_uid = ('V' + uuid.uuid4().hex)[:32]

        utilisateur = Utilisateur.objects.create(
            uid_carte=raw_uid,
            nom=data['nom'],
            prenom=data['prenom'],
            date_naissance=data['date_naissance'],
            lieu_naissance=data['lieu_naissance'],
            departement_pays=data.get('departement_pays', ''),
            nationalite=data['nationalite'],
            societe_raison=data.get('societe_raison', ''),
            statut_validation='EN_ATTENTE'  # Nouveau statut par défaut
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {utilisateur.prenom} {utilisateur.nom} enregistré en attente de validation',
            'user_id': utilisateur.id
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'enregistrement: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["POST"])
def valider_utilisateur(request, user_id):
    """Valider un utilisateur en attente (admin uniquement)"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé - Administrateur requis'})
    
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        if utilisateur.statut_validation != 'EN_ATTENTE':
            return JsonResponse({'success': False, 'message': 'Cet utilisateur n\'est pas en attente de validation'})
        
        utilisateur.statut_validation = 'VALIDE'
        utilisateur.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {utilisateur.prenom} {utilisateur.nom} validé avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la validation: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["POST"])
def refuser_utilisateur(request, user_id):
    """Refuser un utilisateur en attente et le blacklister (admin uniquement)"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé - Administrateur requis'})
    
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        if utilisateur.statut_validation != 'EN_ATTENTE':
            return JsonResponse({'success': False, 'message': 'Cet utilisateur n\'est pas en attente de validation'})
        
        utilisateur.statut_validation = 'VALIDE'
        utilisateur.blackliste = True
        utilisateur.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {utilisateur.prenom} {utilisateur.nom} refusé et blacklisté'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du refus: {str(e)}'
        })

def _import_excel_core(request, statut_validation='VALIDE'):
    """Logique commune d'import Excel.
    statut_validation : 'VALIDE' (import direct) ou 'EN_ATTENTE' (import en attente de validation).
    En-têtes attendus (insensibles à la casse) :
      uid_carte (optionnel), nom, prenom, date_naissance, lieu_naissance,
      departement_pays, nationalite, societe_raison
    """
    file = request.FILES.get('file')
    server_path = request.POST.get('path', '').strip()
    if not file and not server_path:
        return JsonResponse({'success': False, 'message': 'Aucun fichier fourni'})

    if file:
        wb = load_workbook(filename=file, data_only=True)
    else:
        if not os.path.exists(server_path):
            return JsonResponse({'success': False, 'message': 'Chemin de fichier introuvable'})
        wb = load_workbook(filename=server_path, data_only=True)

    ws = wb.active
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = str(cell.value or '').strip().lower()
        if key:
            headers[key] = idx

    required = ['nom', 'prenom', 'date_naissance', 'lieu_naissance', 'departement_pays', 'nationalite', 'societe_raison']
    missing = [h for h in required if h not in headers]
    if missing:
        return JsonResponse({'success': False, 'message': f'Colonnes manquantes: {", ".join(missing)}'})

    def parse_date(val):
        if val is None:
            raise ValueError('Date manquante')
        if isinstance(val, datetime):
            return val.date().isoformat()
        s = str(val).strip()
        if not s:
            raise ValueError('Date vide')
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                pass
        raise ValueError(f'Date invalide: {s}')

    created = 0
    updated = 0
    errors = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        try:
            get = lambda name: r[headers[name]-1] if name in headers else None
            uid_raw = str(get('uid_carte') or '').strip()
            if not uid_raw:
                uid_raw = ('I' + uuid.uuid4().hex)[:32]
            uid_raw = uid_raw.upper()
            payload = {
                'uid_carte': uid_raw,
                'nom': str(get('nom') or '').strip(),
                'prenom': str(get('prenom') or '').strip(),
                'date_naissance': parse_date(get('date_naissance')),
                'lieu_naissance': str(get('lieu_naissance') or '').strip(),
                'departement_pays': str(get('departement_pays') or '').strip(),
                'nationalite': str(get('nationalite') or '').strip(),
                'societe_raison': str(get('societe_raison') or '').strip(),
                'statut_validation': statut_validation,
            }
            try:
                with transaction.atomic():
                    u = Utilisateur.objects.select_for_update().get(uid_carte=uid_raw)
                    for k, v in payload.items():
                        if k != 'uid_carte':
                            setattr(u, k, v)
                    u.save()
                    updated += 1
            except Utilisateur.DoesNotExist:
                try:
                    with transaction.atomic():
                        Utilisateur.objects.create(**payload)
                        created += 1
                except IntegrityError:
                    updated += 1
        except Exception as e:
            errors.append(str(e))

    label = "en attente" if statut_validation == 'EN_ATTENTE' else "terminé"
    return JsonResponse({
        'success': True, 'created': created, 'updated': updated, 'errors': errors,
        'message': f"Import {label}: {created} créé(s), {updated} mis à jour"
    })


@csrf_exempt
@require_http_methods(["POST"])
def import_utilisateurs_excel(request):
    """Importer des utilisateurs validés depuis un fichier Excel."""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé - Administrateur requis'})
    try:
        return _import_excel_core(request, statut_validation='VALIDE')
    except Exception as e:
        logger.error(f"Import Excel échoué: {e}")
        return JsonResponse({'success': False, 'message': f'Erreur import: {e}'})


@csrf_exempt
@require_http_methods(["POST"])
def import_utilisateurs_attente_excel(request):
    """Importer des utilisateurs en attente de validation depuis un fichier Excel."""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé - Administrateur requis'})
    try:
        return _import_excel_core(request, statut_validation='EN_ATTENTE')
    except Exception as e:
        logger.error(f"Import Excel en attente échoué: {e}")
        return JsonResponse({'success': False, 'message': f'Erreur import: {e}'})

def utilisateurs(request):
    """Page des utilisateurs.
    - Admin: accès complet
    - Non-admin: lecture seule + actions limitées côté interface
    """

    enforce_retention_policies()

    search_query = request.GET.get('search', '').strip()
    
    # Utilisateurs validés
    users_valides_qs = Utilisateur.objects.filter(statut_validation='VALIDE')
    if search_query:
        users_valides_qs = users_valides_qs.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(societe_raison__icontains=search_query) |
            Q(lieu_naissance__icontains=search_query) |
            Q(nationalite__icontains=search_query)
        )
    users_valides_qs = users_valides_qs.order_by('nom', 'prenom')
    
    # Utilisateurs en attente de validation
    users_attente_qs = Utilisateur.objects.filter(statut_validation='EN_ATTENTE')
    if search_query:
        users_attente_qs = users_attente_qs.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(societe_raison__icontains=search_query) |
            Q(lieu_naissance__icontains=search_query) |
            Q(nationalite__icontains=search_query)
        )
    users_attente_qs = users_attente_qs.order_by('nom', 'prenom')

    return render(request, 'presence/utilisateurs.html', {
        'is_admin': get_admin_status(request),
        'utilisateurs': users_valides_qs,
        'utilisateurs_attente': users_attente_qs,
        'search_query': search_query,
    })


def historique(request):
    """Page Historique - liste des entrées/sorties (admin uniquement)."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')

    enforce_retention_policies()

    # Recherche libre sur nom/prénom/entreprise
    search_query = request.GET.get('search', '').strip()
    historique_list = (
        HistoriquePresence.objects.select_related('utilisateur')
        .all()
        .order_by('-horodatage')
    )
    if search_query:
        historique_list = historique_list.filter(
            Q(utilisateur__nom__icontains=search_query)
            | Q(utilisateur__prenom__icontains=search_query)
            | Q(utilisateur__societe_raison__icontains=search_query)
        )

    context = {
        'historique': historique_list,
        'search_query': search_query,
        'is_admin': True,
    }
    return render(request, 'presence/historique.html', context)


def statistiques(request):
    """Page dédiée aux statistiques (admin uniquement)."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    # Rétention
    enforce_retention_policies()

    selected_entreprise = request.GET.get('entreprise', '').strip()
    historique_list = HistoriquePresence.objects.select_related('utilisateur').all()
    if selected_entreprise:
        historique_list = historique_list.filter(utilisateur__societe_raison=selected_entreprise)

    now = timezone.localtime()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    # Fenêtre glissante: dernières 24 heures, avec l'heure actuelle à DROITE
    now_floor = now.replace(minute=0, second=0, microsecond=0)
    window_start = now_floor - timezone.timedelta(hours=23)

    per_hour_entrees = [0]*24
    per_hour_sorties = [0]*24
    for i in range(24):
        h_start = window_start + timezone.timedelta(hours=i)
        h_end = h_start + timezone.timedelta(hours=1)
        per_hour_entrees[i] = historique_list.filter(type_action='ENTREE', horodatage__gte=h_start, horodatage__lt=h_end).count()
        per_hour_sorties[i] = historique_list.filter(type_action='SORTIE', horodatage__gte=h_start, horodatage__lt=h_end).count()

    # Courbe Présents cumulés (24h glissantes):
    # Baseline = présents au début de fenêtre (window_start)
    users_qs = Utilisateur.objects.all()
    if selected_entreprise:
        users_qs = users_qs.filter(societe_raison=selected_entreprise)
    current_present = users_qs.filter(present=True).count()
    moves_ent = historique_list.filter(type_action='ENTREE', horodatage__gte=window_start, horodatage__lt=now).count()
    moves_sort = historique_list.filter(type_action='SORTIE', horodatage__gte=window_start, horodatage__lt=now).count()
    baseline = current_present - (moves_ent - moves_sort)
    if baseline < 0:
        baseline = 0

    presents_cum = []
    current = baseline
    for i in range(24):
        current += per_hour_entrees[i]
        current -= per_hour_sorties[i]
        if current < 0:
            current = 0
        presents_cum.append(current)

    # 7 derniers jours (totaux / jours) + Heatmap heures×jours
    last7_labels = []
    last7_entrees = []
    last7_sorties = []
    # Heatmap données détaillées (entrées, sorties, total)
    heatmap_days = []
    heatmap_data = []  # liste d'objets {x: dayIndex, y: hour, e: nb_entrees, s: nb_sorties, t: total}
    for i in range(6, -1, -1):
        day = today_start - timezone.timedelta(days=i)
        next_day = day + timezone.timedelta(days=1)
        last7_labels.append(day.strftime('%d/%m'))
        # Totaux par jour
        last7_entrees.append(historique_list.filter(type_action='ENTREE', horodatage__gte=day, horodatage__lt=next_day).count())
        last7_sorties.append(historique_list.filter(type_action='SORTIE', horodatage__gte=day, horodatage__lt=next_day).count())
        # Label pour heatmap (jour de semaine + date)
        heatmap_days.append(day.strftime('%a %d/%m'))
        # Remplir 24 heures
        for h in range(24):
            h_start = day + timezone.timedelta(hours=h)
            h_end = h_start + timezone.timedelta(hours=1)
            e = historique_list.filter(type_action='ENTREE', horodatage__gte=h_start, horodatage__lt=h_end).count()
            s = historique_list.filter(type_action='SORTIE', horodatage__gte=h_start, horodatage__lt=h_end).count()
            heatmap_data.append({'x': len(heatmap_days)-1, 'y': h, 'e': e, 's': s, 't': e + s})

    entreprises = list(
        HistoriquePresence.objects.select_related('utilisateur')
        .values_list('utilisateur__societe_raison', flat=True)
        .distinct().order_by('utilisateur__societe_raison')
    )

    stats = {
        'per_hour_entrees': per_hour_entrees,
        'per_hour_sorties': per_hour_sorties,
        'presents_cum': presents_cum,
        'last7_labels': last7_labels,
        'last7_entrees': last7_entrees,
        'last7_sorties': last7_sorties,
        'heatmap_days': heatmap_days,
        'heatmap_data': heatmap_data,
    }

    context = {
        'is_admin': True,
        'entreprises': entreprises,
        'selected_entreprise': selected_entreprise,
        'stats': stats,
    }
    return render(request, 'presence/statistiques.html', context)


def _tail(filepath, max_lines=500):
    """Lire les dernières lignes d'un fichier de logs sans tout charger en mémoire."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.readlines()[-max_lines:]
    except FileNotFoundError:
        return ["Aucun fichier de logs trouvé."]
    except Exception as e:
        return [f"Erreur lecture logs: {e}"]


def logs_page(request):
    """Page de visualisation des logs applicatifs (admin uniquement)."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    enforce_retention_policies()

    level = request.GET.get('level', '').upper().strip()
    query = request.GET.get('q', '').strip()

    from django.conf import settings
    log_file = (settings.BASE_DIR / 'logs' / 'app.log')
    lines = _tail(log_file, max_lines=1000)

    # Filtrage basique par niveau et requête texte + parsing
    import re
    pattern = re.compile(r"^(?P<ts>[^\[]+)\s\[(?P<lvl>[A-Z]+)\]\s(?P<logger>[^:]+):\s(?P<msg>.*)$")
    entries = []
    for raw in lines:
        line = raw.rstrip('\n')
        if level and f"[{level}]" not in line:
            continue
        if query and query.lower() not in line.lower():
            continue
        m = pattern.match(line)
        if m:
            entries.append({
                'timestamp': m.group('ts').strip(),
                'level': m.group('lvl'),
                'logger': m.group('logger').strip(),
                'message': m.group('msg'),
                'raw': line,
            })
        else:
            entries.append({
                'timestamp': '',
                'level': 'INFO',
                'logger': '',
                'message': line,
                'raw': line,
            })

    context = {
        'log_entries': entries[::-1],  # les plus récents d'abord
        'level': level,
        'q': query,
        'is_admin': True,
    }
    return render(request, 'presence/logs.html', context)

@require_http_methods(["POST"])
def clear_logs(request):
    """Vider le fichier de logs (admin uniquement)."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')

    from django.conf import settings
    log_file = (settings.BASE_DIR / 'logs' / 'app.log')
    try:
        # Tronquer le fichier
        with open(log_file, 'w', encoding='utf-8'):
            pass
        messages.success(request, 'Logs supprimés')
    except Exception as e:
        messages.error(request, f'Impossible de supprimer les logs: {e}')
    return redirect('logs_page')

@require_http_methods(["POST"])
def purge_now(request):
    """Déclencher immédiatement la purge selon les règles de rétention (admin uniquement)."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    try:
        enforce_retention_policies()
        messages.success(request, "Purge des anciennes données appliquée (historique/logs)")
    except Exception as e:
        messages.error(request, f"Erreur lors de la purge: {e}")
    return redirect('parametres')

def parametres(request):
    """Page Paramètres: mot de passe admin + rétention des données."""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')

    enforce_retention_policies()

    if request.method == 'POST':
        action = request.POST.get('action', 'update_password')
        if action == 'update_password':
            old = request.POST.get('old_password', '')
            new1 = request.POST.get('new_password1', '')
            new2 = request.POST.get('new_password2', '')
            try:
                cfg = ConfigurationSession.objects.get(cle='admin_password')
                stored_hash = cfg.valeur
            except ConfigurationSession.DoesNotExist:
                stored_hash = make_password('admin123')
                cfg = ConfigurationSession.objects.create(cle='admin_password', valeur=stored_hash)

            if not check_password(old, stored_hash):
                messages.error(request, "Ancien mot de passe incorrect")
            elif len(new1) < 6:
                messages.error(request, "Le nouveau mot de passe doit contenir au moins 6 caractères")
            elif new1 != new2:
                messages.error(request, "Les deux champs du nouveau mot de passe ne correspondent pas")
            else:
                cfg.valeur = make_password(new1)
                cfg.save()
                messages.success(request, "Mot de passe mis à jour")
        elif action == 'update_retention':
            # Mettre à jour les durées de rétention (jours)
            hist = request.POST.get('retention_history_days', '90')
            logs = request.POST.get('retention_logs_days', '90')
            users_enabled = request.POST.get('retention_users_inactive_enabled', '0')
            users_days = request.POST.get('retention_users_inactive_days', '0')
            try:
                hist_days = max(1, int(hist))
                logs_days = max(1, int(logs))
                users_days_int = max(0, int(users_days))
                _set_cfg('retention_history_days', str(hist_days))
                _set_cfg('retention_logs_days', str(logs_days))
                _set_cfg('retention_users_inactive_enabled', '1' if users_enabled in ('1', 'on', 'true', 'True') else '0')
                _set_cfg('retention_users_inactive_days', str(users_days_int))
                messages.success(request, f"Rétention mise à jour (Historique: {hist_days}j, Logs: {logs_days}j, Utilisateurs inactifs: {'ON' if users_enabled in ('1','on','true','True') else 'OFF'}{f'/{users_days_int}j' if users_enabled in ('1','on','true','True') else ''})")
            except ValueError:
                messages.error(request, "Valeurs de rétention invalides")

    # Valeurs actuelles pour affichage
    try:
        hist_days = int(_get_cfg('retention_history_days', '90'))
    except ValueError:
        hist_days = 90
    try:
        logs_days = int(_get_cfg('retention_logs_days', '90'))
    except ValueError:
        logs_days = 90
    users_inactive_enabled = _get_cfg('retention_users_inactive_enabled', '0') in ('1','true','True')
    try:
        users_inactive_days = int(_get_cfg('retention_users_inactive_days', '0'))
    except ValueError:
        users_inactive_days = 0

    return render(request, 'presence/parametres.html', {
        'is_admin': True,
        'retention_history_days': hist_days,
        'retention_logs_days': logs_days,
        'retention_users_inactive_enabled': users_inactive_enabled,
        'retention_users_inactive_days': users_inactive_days,
    })

@require_http_methods(["POST"])
def supprimer_utilisateur(request, user_id):
    """Supprimer un utilisateur"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        nom_complet = f"{utilisateur.prenom} {utilisateur.nom}"
        utilisateur.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {nom_complet} supprimé avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["POST"])
def modifier_utilisateur(request, user_id):
    """Modifier un utilisateur"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        data = json.loads(request.body)
        
        utilisateur.nom = data['nom']
        utilisateur.prenom = data['prenom']
        utilisateur.date_naissance = data['date_naissance']
        utilisateur.lieu_naissance = data['lieu_naissance']
        utilisateur.departement_pays = data.get('departement_pays', '')
        utilisateur.nationalite = data['nationalite']
        utilisateur.societe_raison = data.get('societe_raison', '')
        utilisateur.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {utilisateur.prenom} {utilisateur.nom} modifié avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la modification: {str(e)}'
        })

def export_utilisateurs_excel(request):
    """Exporter la liste des utilisateurs validés en Excel"""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs validés"
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Date de naissance', 'Lieu de naissance',
        'Département/Pays', 'Nationalité', 'Société/Raison',
        'Statut', 'Date d\'enregistrement'
    ]
    
    for col, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col)}1'] = header
    
    # Données - uniquement les utilisateurs validés
    utilisateurs_data = Utilisateur.objects.filter(statut_validation='VALIDE').order_by('nom', 'prenom')
    for row, utilisateur in enumerate(utilisateurs_data, 2):
        ws[f'A{row}'] = utilisateur.nom
        ws[f'B{row}'] = utilisateur.prenom
        ws[f'C{row}'] = utilisateur.date_naissance.strftime('%d/%m/%Y')
        ws[f'D{row}'] = utilisateur.lieu_naissance
        ws[f'E{row}'] = utilisateur.departement_pays
        ws[f'F{row}'] = utilisateur.nationalite
        ws[f'G{row}'] = utilisateur.societe_raison
        ws[f'H{row}'] = 'Présent' if utilisateur.present else 'Absent'
        ws[f'I{row}'] = utilisateur.date_creation.strftime('%d/%m/%Y %H:%M')
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="utilisateurs_valides_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

def export_utilisateurs_attente_excel(request):
    """Exporter la liste des utilisateurs en attente en Excel"""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs en attente"
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Date de naissance', 'Lieu de naissance',
        'Département/Pays', 'Nationalité', 'Société/Raison',
        'Date d\'enregistrement'
    ]
    
    for col, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col)}1'] = header
    
    # Données - uniquement les utilisateurs en attente
    utilisateurs_data = Utilisateur.objects.filter(statut_validation='EN_ATTENTE').order_by('nom', 'prenom')
    for row, utilisateur in enumerate(utilisateurs_data, 2):
        ws[f'A{row}'] = utilisateur.nom
        ws[f'B{row}'] = utilisateur.prenom
        ws[f'C{row}'] = utilisateur.date_naissance.strftime('%d/%m/%Y')
        ws[f'D{row}'] = utilisateur.lieu_naissance
        ws[f'E{row}'] = utilisateur.departement_pays
        ws[f'F{row}'] = utilisateur.nationalite
        ws[f'G{row}'] = utilisateur.societe_raison
        ws[f'H{row}'] = utilisateur.date_creation.strftime('%d/%m/%Y %H:%M')
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="utilisateurs_attente_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

def export_historique_excel(request):
    """Exporter l'historique en Excel"""
    if not get_admin_status(request):
        messages.error(request, "Accès refusé - Mode administrateur requis")
        return redirect('accueil')
    
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique des présences"
    
    # En-têtes
    headers = ['Nom', 'Prénom', 'Société/Raison', 'Action', 'Date et heure']
    
    for col, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col)}1'] = header
    
    # Données
    historique_data = HistoriquePresence.objects.select_related('utilisateur').all().order_by('-horodatage')
    for row, entry in enumerate(historique_data, 2):
        ws[f'A{row}'] = entry.utilisateur.nom
        ws[f'B{row}'] = entry.utilisateur.prenom
        ws[f'C{row}'] = entry.utilisateur.societe_raison
        ws[f'D{row}'] = entry.get_type_action_display()
        ws[f'E{row}'] = entry.horodatage.strftime('%d/%m/%Y %H:%M:%S')
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="historique_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

@require_http_methods(["POST"])
def vider_liste_presences(request):
    """Vider la liste des personnes présentes"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        nombre_modifies = Utilisateur.objects.filter(present=True).update(present=False)
        return JsonResponse({
            'success': True,
            'message': f'{nombre_modifies} personnes marquées comme absentes'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["POST"])
def forcer_sortie_utilisateur(request):
    """Forcer la sortie d'un utilisateur spécifique (admin uniquement)"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé - Administrateur requis'})
    
    try:
        data = json.loads(request.body or '{}')
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'success': False, 'message': 'ID utilisateur manquant'})
        
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        
        if not utilisateur.present:
            return JsonResponse({
                'success': False,
                'message': f'{utilisateur.prenom} {utilisateur.nom} n\'est pas présent(e)'
            })
        
        # Marquer la sortie
        utilisateur.present = False
        utilisateur.save()
        HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='SORTIE')
        
        return JsonResponse({
            'success': True,
            'message': f'Sortie forcée: {utilisateur.prenom} {utilisateur.nom}',
            'utilisateur': {
                'nom': utilisateur.nom,
                'prenom': utilisateur.prenom,
                'societe_raison': utilisateur.societe_raison
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@require_http_methods(["POST"])
def supprimer_tous_utilisateurs(request):
    """Supprimer tous les utilisateurs"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        nombre_supprimes = Utilisateur.objects.count()
        Utilisateur.objects.all().delete()
        return JsonResponse({
            'success': True,
            'message': f'{nombre_supprimes} utilisateurs supprimés'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@require_http_methods(["POST"])
def vider_historique(request):
    """Vider l'historique"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        nombre_supprimes = HistoriquePresence.objects.count()
        HistoriquePresence.objects.all().delete()
        return JsonResponse({
            'success': True,
            'message': f'{nombre_supprimes} entrées d\'historique supprimées'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@require_http_methods(["GET"])
def get_utilisateur_data(request, user_id):
    """Récupérer les données d'un utilisateur pour modification"""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        return JsonResponse({
            'success': True,
            'data': {
                'nom': utilisateur.nom,
                'prenom': utilisateur.prenom,
                'date_naissance': utilisateur.date_naissance.strftime('%Y-%m-%d'),
                'lieu_naissance': utilisateur.lieu_naissance,
                'departement_pays': utilisateur.departement_pays,
                'nationalite': utilisateur.nationalite,
                'societe_raison': utilisateur.societe_raison,
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["GET"])
def ecoute_nfc_continue(request):
    """Endpoint pour l'écoute continue des cartes NFC.
    Ne déclenche une action qu'à l'INsertion d'une carte (frontière no-card -> card).
    Tant que la même carte reste posée, aucune nouvelle action n'est générée.
    """
    global nfc_service, last_card_detection, last_seen_uid
    
    try:
        if nfc_service is None:
            nfc_service = get_nfc_service()
        
        # Vérifier l'état du lecteur
        reader_ok, msg = nfc_service.test_connection()
        
        # Si le lecteur est déconnecté, réinitialiser l'état et retourner immédiatement
        if not reader_ok:
            last_seen_uid = None
            return JsonResponse({
                'success': False, 
                'no_card': True, 
                'reader_connected': False,
                'card_present': False,
                'message': msg or 'Lecteur déconnecté'
            })
        
        # Lire une carte (non bloquant). Retourne None s'il n'y a pas de carte.
        uid = nfc_service.lire_carte()

        if uid:
            current_time = time.time()

            # Si la même carte est toujours posée, ne rien faire (edge-only)
            if last_seen_uid == uid:
                return JsonResponse({
                    'success': False,
                    'no_card': True,
                    'reader_connected': True,
                    'card_present': True,
                    'message': 'Même carte toujours présente, en attente de retrait'
                })

            # Nouvelle carte détectée (transition no-card -> card)
            last_seen_uid = uid
            return JsonResponse(_resoudre_uid(uid))
        else:
            # Pas de carte détectée - réinitialiser l'état edge et retourner état d'attente
            if last_seen_uid is not None:
                last_seen_uid = None
            return JsonResponse({
                'success': False,
                'no_card': True,
                'reader_connected': True,
                'card_present': False,
                'message': 'En attente d\'une carte NFC'
            })
            
    except Exception as e:
        logger.error(f"Erreur dans ecoute_nfc_continue: {e}")
        return JsonResponse({
            'success': False,
            'error': True,
            'reader_connected': False,
            'message': f'Erreur de lecture NFC: {str(e)}'
        })

@require_http_methods(["POST"])
def toggle_blacklist(request, user_id):
    """Basculer l'état blacklisté d'un utilisateur (admin uniquement)."""
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        utilisateur.blackliste = not getattr(utilisateur, 'blackliste', False)
        utilisateur.save()
        state = 'blacklisté' if utilisateur.blackliste else 'retiré de la blacklist'
        return JsonResponse({'success': True, 'message': f"{utilisateur.prenom} {utilisateur.nom} {state}", 'blackliste': utilisateur.blackliste})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["POST"])
def detacher_carte(request, user_id):
    """Détacher la carte personnelle d'un utilisateur (génère un UID synthétique).
    Admin uniquement. Utile si un utilisateur a été créé avec l'UID d'un badge visiteur.
    """
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        # Générer un UID unique synthétique
        new_uid = None
        for _ in range(5):
            candidate = ('V' + uuid.uuid4().hex)[:32]
            if not Utilisateur.objects.filter(uid_carte=candidate).exists():
                new_uid = candidate
                break
        if new_uid is None:
            return JsonResponse({'success': False, 'message': "Impossible de générer un UID unique, réessayez."})
        old_uid = utilisateur.uid_carte
        utilisateur.uid_carte = new_uid
        utilisateur.save()
        return JsonResponse({'success': True, 'message': "Carte détachée: l'utilisateur n'est plus lié à l'ancienne carte.", 'old_uid': old_uid})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["POST"])
def lier_carte(request, user_id):
    """Lier une carte (UID) à un utilisateur.
    Conflits gérés:
    - UID déjà lié à un autre utilisateur -> erreur avec nom/prénom
    - UID existant comme BadgeVisiteur -> erreur; si affecté, préciser à qui
    """
    try:
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        data = json.loads(request.body or '{}')
        uid = (data.get('uid') or '').strip().upper()
        if not uid:
            return JsonResponse({'success': False, 'message': 'UID manquant'})

        # Conflit avec un autre utilisateur (case-insensitive)
        other = (
            Utilisateur.objects
            .filter(uid_carte__iexact=uid)
            .exclude(id=utilisateur.id)
            .first()
        )
        if other:
            return JsonResponse({'success': False, 'code': 'ALREADY_LINKED', 'message': f"Badge déjà lié à {other.prenom} {other.nom}"})

        # Conflit avec un badge visiteur (bloquant)
        b = BadgeVisiteur.objects.select_related('affecte_a').filter(uid_carte__iexact=uid).first()
        if b:
            if b.affecte_a:
                return JsonResponse({'success': False, 'code': 'VISITOR_ASSIGNED', 'message': f"Badge déjà lié à {b.affecte_a.prenom} {b.affecte_a.nom}"})
            return JsonResponse({'success': False, 'code': 'VISITOR_BADGE', 'message': 'UID correspond à un badge visiteur non affecté'})

        # Lier l'UID à l'utilisateur courant, transaction + contrainte DB
        try:
            with transaction.atomic():
                ulock = Utilisateur.objects.select_for_update().get(id=user_id)
                ulock.uid_carte = uid
                ulock.save()
        except IntegrityError:
            # Contrainte unique déclenchée par un autre utilisateur
            other = (
                Utilisateur.objects
                .filter(uid_carte__iexact=uid)
                .exclude(id=user_id)
                .first()
            )
            if other:
                return JsonResponse({'success': False, 'code': 'ALREADY_LINKED', 'message': f"Badge déjà lié à {other.prenom} {other.nom}"})
            return JsonResponse({'success': False, 'message': 'Badge déjà lié à un autre utilisateur'})

        return JsonResponse({'success': True, 'message': 'Badge lié avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


# -------------------- Gestion Badges Visiteurs (Admin) --------------------
def visiteurs(request):
    """Page de gestion des badges visiteurs.
    - Admin: toutes actions
    - Non-admin: uniquement désaffecter (délier)
    """
    enforce_retention_policies()
    badges = BadgeVisiteur.objects.select_related('affecte_a').all().order_by('uid_carte')
    return render(request, 'presence/visiteurs.html', {
        'is_admin': get_admin_status(request),
        'badges': badges,
    })


@require_http_methods(["GET"])  # simple lecture UID sans effet côté modèles
def nfc_lire_uid(request):
    """Lire l'UID d'une carte NFC et retourner son statut d'affectation.
    status:
      - user_linked: lié à un Utilisateur
      - visitor_assigned: badge visiteur affecté
      - visitor_free: badge visiteur libre
      - unknown: non référencé
    """
    global nfc_service
    try:
        if nfc_service is None:
            nfc_service = get_nfc_service()
        reader_ok, _msg = nfc_service.test_connection()
        if not reader_ok:
            return JsonResponse({'success': False, 'reader_connected': False, 'message': 'Lecteur déconnecté'})
        uid = nfc_service.lire_carte()
        if uid:
            # Déterminer le statut
            linked_user = Utilisateur.objects.filter(uid_carte__iexact=uid).first()
            if linked_user:
                return JsonResponse({
                    'success': True,
                    'uid': uid,
                    'reader_connected': True,
                    'status': 'user_linked',
                    'linked_user': {
                        'id': linked_user.id,
                        'nom': linked_user.nom,
                        'prenom': linked_user.prenom
                    }
                })
            badge = BadgeVisiteur.objects.select_related('affecte_a').filter(uid_carte__iexact=uid).first()
            if badge:
                if badge.affecte_a:
                    return JsonResponse({
                        'success': True,
                        'uid': uid,
                        'reader_connected': True,
                        'status': 'visitor_assigned',
                        'linked_user': {
                            'id': badge.affecte_a.id,
                            'nom': badge.affecte_a.nom,
                            'prenom': badge.affecte_a.prenom
                        }
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'uid': uid,
                        'reader_connected': True,
                        'status': 'visitor_free'
                    })
            # Inconnu
            return JsonResponse({'success': True, 'uid': uid, 'reader_connected': True, 'status': 'unknown'})
        return JsonResponse({'success': False, 'uid': None, 'reader_connected': True, 'message': 'Aucune carte détectée'})
    except Exception as e:
        logger.debug(f"nfc_lire_uid error: {e}")
        return JsonResponse({'success': False, 'reader_connected': False, 'message': str(e)})


@require_http_methods(["POST"])
def visiteur_ajouter_badge(request):
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    try:
        data = json.loads(request.body or '{}')
        uid = data.get('uid', '').strip()
        nom = (data.get('nom') or '').strip()
        if not uid:
            return JsonResponse({'success': False, 'message': 'UID manquant'})
        badge, created = BadgeVisiteur.objects.get_or_create(uid_carte=uid)
        # Mettre à jour le nom si fourni (à la création ou si vide)
        if nom:
            if created or not badge.nom:
                badge.nom = nom
                badge.save()
        if created:
            return JsonResponse({'success': True, 'message': 'Badge visiteur ajouté', 'badge_id': badge.id})
        return JsonResponse({'success': True, 'message': 'Badge visiteur déjà existant', 'badge_id': badge.id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["POST"])
def visiteur_supprimer_badge(request, badge_id):
    if not get_admin_status(request):
        return JsonResponse({'success': False, 'message': 'Accès refusé'})
    try:
        badge = get_object_or_404(BadgeVisiteur, id=badge_id)
        badge.delete()
        return JsonResponse({'success': True, 'message': 'Badge visiteur supprimé'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["POST"])
def visiteur_desaffecter_badge(request, badge_id):
    # Ouvert aux non-admins (délier) selon besoin opérationnel
    try:
        badge = get_object_or_404(BadgeVisiteur, id=badge_id)
        badge.affecte_a = None
        badge.date_attribution = None
        badge.save()
        return JsonResponse({'success': True, 'message': 'Badge désaffecté'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["POST"])
def visiteur_affecter_badge(request):
    """Affecter un badge visiteur à un utilisateur, et le marquer ENTRÉE immédiatement.
    Ouvert aux non-admins pour permettre l'affectation depuis l'accueil.
    """
    try:
        data = json.loads(request.body or '{}')
        uid = data.get('uid')
        badge_id = data.get('badge_id')
        user_id = data.get('user_id')
        if not user_id or (not uid and not badge_id):
            return JsonResponse({'success': False, 'message': 'Paramètres manquants'})
        if uid:
            badge = get_object_or_404(BadgeVisiteur, uid_carte=uid)
        else:
            badge = get_object_or_404(BadgeVisiteur, id=badge_id)
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        # Ne pas permettre l'affectation d'un utilisateur blacklisté
        if getattr(utilisateur, 'blackliste', False):
            return JsonResponse({'success': False, 'code': 'BLACKLIST', 'message': f"Utilisateur blacklisté: {utilisateur.prenom} {utilisateur.nom}. Affectation interdite."})
        # Ne pas permettre l'affectation d'un utilisateur en attente de validation
        if getattr(utilisateur, 'statut_validation', 'VALIDE') == 'EN_ATTENTE':
            return JsonResponse({'success': False, 'code': 'EN_ATTENTE', 'message': f"Utilisateur en attente de validation: {utilisateur.prenom} {utilisateur.nom}. Affectation interdite."})
        badge.affecte_a = utilisateur
        badge.date_attribution = timezone.now()
        badge.save()
        # Marquer entrée
        if not utilisateur.present:
            utilisateur.present = True
            utilisateur.save()
            HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='ENTREE')
        return JsonResponse({'success': True, 'message': 'Badge affecté', 'badge_id': badge.id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})


@require_http_methods(["GET"])  # Liste courte pour sélecteur d'affectation (ouverte)
def utilisateurs_liste(request):
    q = request.GET.get('q', '').strip()
    # Ne retourner que les utilisateurs validés
    qs = Utilisateur.objects.filter(statut_validation='VALIDE').order_by('nom', 'prenom')
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(societe_raison__icontains=q))
    # Retourner tous les utilisateurs (pas de limite)
    data = [
        {
            'id': u.id,
            'label': f"{u.prenom} {u.nom} - {u.societe_raison}",
            'nom': u.nom,
            'prenom': u.prenom,
            'societe_raison': u.societe_raison,
            'blackliste': bool(getattr(u, 'blackliste', False)),
        } for u in qs
    ]
    return JsonResponse({'success': True, 'results': data})


@csrf_exempt
@require_http_methods(["POST"])
def lier_badge_utilisateur_permanent(request):
    """Lier de façon permanente un badge inconnu à un utilisateur existant.
    Contrairement aux badges visiteurs, cette liaison est permanente et transforme
    le badge en carte personnelle de l'utilisateur.
    """
    try:
        data = json.loads(request.body or '{}')
        uid = data.get('uid')
        user_id = data.get('user_id')
        
        if not uid or not user_id:
            return JsonResponse({'success': False, 'message': 'Paramètres manquants'})
        
        # Vérifier que le badge n'est pas déjà utilisé
        if Utilisateur.objects.filter(uid_carte__iexact=uid).exists():
            return JsonResponse({'success': False, 'message': 'Ce badge est déjà lié à un utilisateur'})
        
        # Vérifier que l'utilisateur existe et est validé
        utilisateur = get_object_or_404(Utilisateur, id=user_id)
        
        if getattr(utilisateur, 'blackliste', False):
            return JsonResponse({
                'success': False,
                'code': 'BLACKLIST',
                'message': f"Utilisateur blacklisté: {utilisateur.prenom} {utilisateur.nom}. Liaison interdite."
            })
        
        if getattr(utilisateur, 'statut_validation', 'VALIDE') == 'EN_ATTENTE':
            return JsonResponse({
                'success': False,
                'code': 'EN_ATTENTE',
                'message': f"Utilisateur en attente de validation: {utilisateur.prenom} {utilisateur.nom}. Liaison interdite."
            })
        
        # Vérifier si l'utilisateur a déjà une carte
        old_uid = utilisateur.uid_carte
        if old_uid and not old_uid.startswith('V') and not old_uid.startswith('I'):
            # L'utilisateur a déjà une vraie carte (pas un UID généré)
            return JsonResponse({
                'success': False,
                'message': f"Cet utilisateur possède déjà une carte (UID: {old_uid}). Détachez d'abord l'ancienne carte."
            })
        
        # Lier le badge à l'utilisateur
        utilisateur.uid_carte = uid.upper()
        utilisateur.save()
        
        # Marquer l'entrée immédiatement
        if not utilisateur.present:
            utilisateur.present = True
            utilisateur.save()
            HistoriquePresence.objects.create(utilisateur=utilisateur, type_action='ENTREE')
        
        return JsonResponse({
            'success': True,
            'message': f'Badge lié à {utilisateur.prenom} {utilisateur.nom}',
            'utilisateur': {
                'nom': utilisateur.nom,
                'prenom': utilisateur.prenom,
                'societe_raison': utilisateur.societe_raison
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {e}'})

